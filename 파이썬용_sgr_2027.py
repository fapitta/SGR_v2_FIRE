import pandas as pd
import re
import numpy as np
import warnings
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, session
from functools import wraps
import io
import os
import datetime
from openpyxl import load_workbook
from scipy.optimize import minimize

# AI 최적화 모듈 import
try:
    from ai_optimizer import AIOptimizationEngine
    AI_MODULE_AVAILABLE = True
except ImportError:
    print("[WARNING] AI optimizer module not available")
    AI_MODULE_AVAILABLE = False

# [OPTIMIZATION] Simple Cache for Google Sheets data
_gsheet_cache = None
_gsheet_cache_time = None

# 경고 무시 설정
warnings.filterwarnings('ignore')

# ----------------------------------------------------------------------
# 0. Secrets Management (Streamlit Cloud Compatibility)
# ----------------------------------------------------------------------
def get_secret(key, default=None):
    """
    스트림릿 클라우드(환경 변수) 혹은 로컬(.streamlit/secrets.toml)에서 정보를 읽어옴
    """
    # 1. 환경 변수 확인 (스트림릿 클라우드용)
    env_val = os.environ.get(key.replace('.', '_').upper())
    if env_val: return env_val

    # 2. 로컬 secrets.toml 확인 (로컬 개발용)
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        secrets_path = os.path.join(current_dir, '.streamlit', 'secrets.toml')
        if os.path.exists(secrets_path):
            with open(secrets_path, 'r', encoding='utf-8') as f:
                content = f.read()
                # 간단한 TOML 파싱 (섹션 및 키-값)
                current_section = None
                lines = content.split('\n')
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('#'): continue
                    if line.startswith('[') and line.endswith(']'):
                        current_section = line[1:-1]
                        continue
                    if '=' in line:
                        k, v = [x.strip() for x in line.split('=', 1)]
                        # 따옴표 제거
                        v = v.strip('"\'')
                        full_key = f"{current_section}.{k}" if current_section else k
                        if full_key == key:
                            return v
    except Exception as e:
        print(f"[WARN] Secrets loading failed: {e}")
    
    return default

def sanitize_data(data):
    """JSON 직렬화 가능하도록 데이터 정제 (NaN/Inf -> None, HTML 정제)"""
    import re
    if isinstance(data, dict):
        return {str(k): sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(v) for v in data]
    elif isinstance(data, float):
        if np.isnan(data) or np.isinf(data): return None
        return float(data)
    elif isinstance(data, str):
        # Extremely aggressive HTML stripping
        return re.sub(r'<[^>]*>', '', data).strip()
    elif hasattr(data, 'tolist'): # Handle numpy arrays
        return sanitize_data(data.tolist())
    return data

# ----------------------------------------------------------------------
# 1. 데이터 로드 및 전처리 클래스 (Advanced DataProcessor)
# ----------------------------------------------------------------------

class DataProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.GROUP_MAPPING = {
            '병원(계)': ['상급종합', '종합병원', '병원', '요양병원'], '의원(계)': ['의원'],
            '치과(계)': ['치과병원', '치과의원'], '한방(계)': ['한방병원', '한의원'], '약국(계)': ['약국']
        }
        self.raw_data = self._load_data()
        self.HOSPITAL_TYPES = ['상급종합', '종합병원', '병원', '요양병원', '의원', '치과병원', '치과의원', '한방병원', '한의원', '약국']

    def _load_sheet(self, sheet_name, index_col=0, filter_years=False):
        try:
            sheet_id = get_secret('google_sheets.sheet_id', '1UNahJhA6bSOJMahQJWFCJzUV33VvjVNafmu2lHRb-sM')
            xlsx_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
            df = pd.read_excel(xlsx_url, sheet_name=sheet_name, index_col=None)
            
            if filter_years:
                target_col = None
                if '연도' in df.columns: target_col = '연도'
                elif 'Year' in df.columns: target_col = 'Year'
                else: target_col = df.columns[0]
                
                df = df.set_index(target_col)
                df.index = pd.to_numeric(df.index, errors='coerce')
                df = df[df.index.notna()]
                df.index = df.index.astype(int)
                df = df[df.index > 1990]
            else:
                if not df.empty and len(df.columns) > 0:
                     df = df.set_index(df.columns[0])

            df.columns = [str(c).strip() for c in df.columns]
            if df.index.dtype == 'object':
                df.index = [str(i).strip() if isinstance(i, str) else i for i in df.index]

            return df.apply(pd.to_numeric, errors='coerce')
        except Exception as e:
            print(f"Sheet {sheet_name} load warning: {e}")
            return pd.DataFrame()

    def _load_data(self, force=False):
        global _gsheet_cache, _gsheet_cache_time
        now = datetime.datetime.now()
        if not force and _gsheet_cache and _gsheet_cache_time and (now - _gsheet_cache_time).seconds < 300:
            return _gsheet_cache

        try:
            sheet_id = get_secret('google_sheets.sheet_id', '1UNahJhA6bSOJMahQJWFCJzUV33VvjVNafmu2lHRb-sM')
            gsheet_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
            # 우선순위: 구글 시트 URL (사용자 요청: 구글 시트 최신 데이터 반영)
            source = gsheet_url
            print(f"[INFO] Loading data from: {source}")
            with pd.ExcelFile(source) as xls:
                new_data = {
                    'df_expenditure': self._load_sheet_from_xls(xls, 'expenditure_real', filter_years=True),
                    'df_weights': self._load_sheet_from_xls(xls, 'cost_structure').T,
                    'df_raw_mei_inf': self._load_sheet_from_xls(xls, 'factor_pd', filter_years=True),
                    'df_gdp': self._load_sheet_from_xls(xls, 'GDP', filter_years=True),
                    'df_pop': self._load_sheet_from_xls(xls, 'pop', filter_years=True),
                    'df_sgr_reval': self._load_sheet_from_xls(xls, 'cf_t', filter_years=True),
                    'df_sgr_law': self._load_sheet_from_xls(xls, 'law', filter_years=True),
                    'df_rel_value': self._load_sheet_from_xls(xls, 'rvs', filter_years=True),
                    'df_num': self._load_sheet_from_xls(xls, 'num', filter_years=True),
                    'df_contract': self._load_sheet_from_xls(xls, 'contract', filter_years=True),
                    'df_finance': self._load_sheet_from_xls(xls, 'finance', filter_years=True),
                    'df_rate_py': self._load_sheet_from_xls(xls, 'rate_py', filter_years=True) if 'rate_py' in xls.sheet_names else self._load_sheet_from_xls(xls, 'Sheet1', filter_years=True)
                }
            _gsheet_cache = new_data
            _gsheet_cache_time = now
            return new_data
        except Exception as e:
            print(f"❌ 데이터 로드 치명적 오류: {e}")
            return {k: pd.DataFrame() for k in ['df_expenditure','df_weights','df_raw_mei_inf','df_gdp','df_pop','df_sgr_reval','df_sgr_law','df_rel_value','df_num','df_contract','df_finance','df_rate_py']}

    def _load_sheet_from_xls(self, xls, sheet_name, index_col=0, filter_years=False):
        try:
            if sheet_name not in xls.sheet_names: return pd.DataFrame()
            # Explicitly using engine='openpyxl' is already handled by ExcelFile if it's .xlsx
            df = pd.read_excel(xls, sheet_name=sheet_name, index_col=None)
            
            # Global Cleaning Helper
            import re
            def clean_str(v):
                if not isinstance(v, str): return v
                # Remove all HTML tags and strip
                cleaned = re.sub(r'<[^>]*>', '', v).strip()
                # Fix encoding artifacts if any (e.g. )
                return cleaned.encode('utf-8', 'ignore').decode('utf-8')

            if filter_years:
                target_col = None
                for c in ['연도', 'Year', 'year']:
                    if c in df.columns: 
                        target_col = c
                        break
                if not target_col: target_col = df.columns[0]
                
                df = df.set_index(target_col)
                df.index = pd.to_numeric(df.index, errors='coerce')
                df = df[df.index.notna()]
                df.index = df.index.astype(int)
            else:
                if not df.empty and len(df.columns) > 0:
                     df = df.set_index(df.columns[0])

            # Apply cleaning to all headers
            df.columns = [clean_str(c) for c in df.columns]
            if df.index.dtype == 'object':
                df.index = [clean_str(i) for i in df.index]

            return df.apply(pd.to_numeric, errors='coerce')
        except Exception as e:
            print(f"Sheet {sheet_name} load warning: {e}")
            return pd.DataFrame()

    def reload_data(self):
        """Force reload data from Excel file (SAFE RELOAD)"""
        print(f"[INFO] Reloading data from {self.file_path}...")
        try:
            temp_data = self._load_data(force=True)
            self.raw_data = temp_data
            print("[SUCCESS] Data reloaded successfully.")
            return True
        except Exception as e:
            print(f"[ERROR] Failed to reload data (keeping previous state): {e}")
            # Do NOT update self.raw_data if reload fails
            return False

    def save_overrides_to_excel(self, overrides, mode='final'):
        """
        Save user overrides back to the Excel file using openpyxl to preserve formatting.
        overrides: dict of {key: value} where key is like '병원_2025', 'I1_2025', etc.
        """
        try:
            target_path = self.file_path
            if mode == 'temp':
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                target_path = os.path.join(os.path.dirname(self.file_path), f"SGR_data_임시_{timestamp}.xlsx")

            wb = load_workbook(self.file_path)
            
            # Helper to find column index by header name
            def get_col_idx(ws, header_name, header_row=1):
                for cell in ws[header_row]:
                    if cell.value == header_name:
                        return cell.column
                return None

            # Helper to find row index by year (mostly in column 1)
            def get_row_idx_by_year(ws, year, year_col=1):
                for row in ws.iter_rows(min_row=2, max_col=year_col): # Assuming year is in first few columns
                    cell = row[year_col-1]
                    if cell.value == int(year) or str(cell.value) == str(year):
                        return cell.row
                return None

            # Mapping Logic
            # Key prefixes/patterns -> Sheet Name, Column Name mapping
            # This logic needs to match _apply_overrides reversed
            
            updates_count = 0
            
            for key, value in overrides.items():
                try:
                    sheet_name = None
                    row_key = None # Year or Row Name
                    col_name = None
                    
                    if '_' not in key: continue

                    # 1. Weights: WEIGHT_{TYPE}_{COL}
                    if key.startswith('WEIGHT_'):
                        parts = key.split('_') # WEIGHT, 병원, 인건비
                        if len(parts) == 3:
                            sheet_name = 'cost_structure'
                            row_key = parts[1] # 병원
                            col_name = parts[2] # 인건비 (Header)
                            # For weights, rows are Types, Col is Category
                            # This sheet usually has Type in Col A?
                            # Let's check logic: _load_sheet('cost_structure').T
                            # This means original sheet has Types as Columns probably? Or Types as Rows?
                            # If we transposed it (.T), then originally:
                            # If Types are indices in DF, then in Excel they are likely in the first column?
                            # Wait, 'cost_structure' usually is small. 
                            # Let's assume standard format: Row=Type, Col=Component OR Row=Component, Col=Type
                            # self._load_sheet('cost_structure') -> index_col=0. 
                            # If .T makes Types indices, then originally Types were Columns.
                            # So Sheet: Rows=Components(Index), Cols=Types.
                            # So we search Col = matching Type. Row = matching Component ('인건비').
                            pass # Logic handled below specially for weights

                    else:
                        # Standard Time Series: FIELD_YEAR
                        field, year = key.rsplit('_', 1)
                        if not year.isdigit(): continue # Skip if not a year-based key
                        year = int(year)
                        row_key = year
                        
                        # MEI
                        if field in ['I1', 'I2', 'I3', 'M1', 'M2', 'Z1', 'Z2']:
                            sheet_name = 'factor_pd'
                            col_map = {'I':'인건비','M':'관리비','Z':'재료비'}
                            col_name = f"{col_map[field[0]]}_{field[1]}"
                        
                        # GDP
                        elif field == 'GDP':
                            sheet_name = 'GDP'
                            col_name = '실질GDP'
                        elif field == 'POP':
                            sheet_name = 'GDP'
                            col_name = '영안인구'
                        
                        # Population
                        elif field == 'NHI_POP':
                            sheet_name = 'pop'
                            # In pop sheet, we might have basic or aged. 
                            # Front end sends 'NHI_POP' for 'basic'. 
                            # Is there 'aged'? user input only mapped 'basic' to NHI_POP_xxx so far in my js?
                            # Ah, js: if (item.key === 'basic') allOverrides[`NHI_POP_${year}`]...
                            # I didn't verify if I added overrides for 'aged'.
                            # In main.js saveAllToExcelFile loop: if (userData.population[year].basic !== undefined) ...
                            # It seems I only supported basic there.
                            col_name = '건보대상자수'

                        # Law
                        elif field.startswith('LAW_') or field == 'LAW':
                             sheet_name = 'law'
                             # Parse type if present
                             parts = field.split('_') # LAW, 병원
                             if len(parts) == 2:
                                 col_name = parts[1] # htype
                             else:
                                 # Fallback if no type? 
                                 pass

                        # RV
                        elif field.startswith('RV_') or field == 'RV':
                             sheet_name = 'rvs'
                             parts = field.split('_')
                             if len(parts) == 2:
                                 col_name = parts[1] # htype

                        # Medical
                        elif field in self.HOSPITAL_TYPES:
                            sheet_name = 'expenditure_real'
                            col_name = field
                        
                        # CF
                        elif field.startswith('CF_') and field[3:] in self.HOSPITAL_TYPES:
                            sheet_name = 'cf_t'
                            col_name = field[3:]

                    # --- EXECUTE UPDATE ---
                    if sheet_name and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # 1. Find Column
                        target_col_idx = get_col_idx(ws, col_name)
                        
                        # 2. Find Row
                        target_row_idx = None
                        if sheet_name == 'cost_structure':
                             # Special handling for weights
                             # Weights logic: Row=Component('인건비'..), Col=Type ? Or vice versa?
                             # In `_load_sheet`: `df_weights = self._load_sheet('cost_structure').T`
                             # The `.T` implies the excel has Types as Columns (and Components as Rows).
                             # So we look for Col = Type (row_key), Row = Component (col_name)
                             # row_key was parts[1] (Type), col_name was parts[2] (Component)
                             # Excel: Headers are Types?
                             target_col_idx = get_col_idx(ws, row_key) # Type
                             # Find row for component
                             for r in range(1, 20): # Scan first 20 rows
                                 if ws.cell(row=r, column=1).value == col_name:
                                     target_row_idx = r
                                     break
                        else:
                             # Standard Time Series
                             target_row_idx = get_row_idx_by_year(ws, row_key)

                        if target_row_idx and target_col_idx:
                            ws.cell(row=target_row_idx, column=target_col_idx).value = float(value)
                            updates_count += 1
                
                except Exception as loop_e:
                    print(f"Error updating key {key}: {loop_e}")
                    continue
            
            wb.save(target_path)
            print(f"[SUCCESS] Saved {updates_count} changes to {target_path}.")
            
            # Only reload data and return special msg if final save
            if mode == 'final':
                self.reload_data()
                return True, f"{updates_count}개 항목 원본 업데이트 성공"
            else:
                return True, f"임시 파일 저장 성공: {os.path.basename(target_path)}"

        except Exception as e:
            print(f"[ERROR] Save to Excel failed: {e}")
            return False, str(e)

# ----------------------------------------------------------------------
# 2. 통합 산출 엔진 (CalculationEngine)
# ----------------------------------------------------------------------

# ... (MeiCalculator, SgrCalculator skipped for brevity) ...

class CalculationEngine:
    def run_full_analysis(self, target_year=2025):
        # 2014년부터 2028년까지 분석 (사용자 요청: 2010-2013 제외)
        years = range(2014, 2029) 
        print(f"[INFO] Analysis Range (Calculated): {list(years)}")
        history = {
            'years': list(years), 
            'S1': {}, 'S2': {}, 'Link': {}, 'MEI': {}, 'GDP': {}, 
            'UAF_S1': {}, 'UAF_S2': {},
            'SGR_S1_INDEX': {}, 'SGR_S2_INDEX': {},
            'Target_S1': {}, 'Target_S2': {}
        }
        details = {'mei_raw': {}, 'sgr_factors': {}}
        
        # Additional bulk data for 2020-2026 comparison
        bulk_sgr = {
            'reval_rates': {}, 'pop_growth': {}, 'law_changes': {}, 
            'gdp_growth': {}, 'scenario_adjustments': {}
        }

        # Key Mapping for MEI display (Match Frontend Keys)
        key_map = {
            '병원': '병원(계)', '의원': '의원(계)', '치과': '치과(계)', 
            '한방': '한방(계)', '약국': '약국(계)'
        }

        for y in years:
            df_mei = self.mei_calc.calc_mei_index_by_year(y)
            if df_mei is not None:
                # Rename indices to match frontend expected keys (with suffix)
                df_renamed = df_mei.rename(index=key_map)
                details['mei_raw'][y] = (df_renamed.round(4)).to_dict()

# ----------------------------------------------------------------------
# 2. 통합 산출 엔진 (CalculationEngine)
# ----------------------------------------------------------------------

# ----------------------------------------------------------------------
# 2. 산출 엔진 (MeiCalculator, SgrCalculator, CalculationEngine)
# ----------------------------------------------------------------------

class MeiCalculator:
    def __init__(self, data, hospital_types):
        self.data = data
        self.hospital_types = hospital_types
        self._cache = {} # [OPTIMIZATION] Memoization

    def calc_mei_index_by_year(self, target_year):
        if target_year in self._cache:
            return self._cache[target_year]
            
        calc_year = target_year - 2
        df_inf = self.data['df_raw_mei_inf']
        df_weights = self.data['df_weights']
        
        try:
            I_types = [c for c in df_inf.columns if '인건비' in c] # 인건비_1, 2, 3
            # 인건비 지수는 최근 3개년 연평균 증가율 적용
            # 2014년 기준 calc_year=2012년인 경우 2009년 데이터가 필요함.
            # 데이터가 2010년부터 있다면 2013년부터 3년 평균 가능.
            # 2012년의 경우 2010~2012 (2년) 또는 1년치만 사용할 수 있도록 예외처리
            try:
                base_year = calc_year - 3
                if base_year not in df_inf.index:
                    base_year = max(df_inf.index.min(), 2010)
                
                years_diff = calc_year - base_year
                if years_diff <= 0:
                    labor_rates = pd.Series(1.0, index=I_types)
                else:
                    labor_rates = (df_inf.loc[calc_year, I_types] / df_inf.loc[base_year, I_types])**(1/years_diff)
            except:
                labor_rates = pd.Series(1.0, index=I_types)
            
            non_I_types = [c for c in df_inf.columns if '관리비' in c or '재료비' in c]
            try:
                raw_rates = df_inf.loc[calc_year, non_I_types] / df_inf.loc[calc_year-1, non_I_types]
            except:
                raw_rates = pd.Series(1.0, index=non_I_types)
            
            M_types = [c for c in raw_rates.index if '관리비' in c]
            Z_types = [c for c in raw_rates.index if '재료비' in c]
            
            results = {}
            for i_t in I_types:
                for m_t in M_types:
                    for z_t in Z_types:
                        name = f"I{i_t.split('_')[-1]}M{m_t.split('_')[-1]}Z{z_t.split('_')[-1]}"
                        # Ensure we only use the specified hospital types and match the names exactly
                        w_l = df_weights['인건비'].reindex(self.hospital_types).fillna(0)
                        w_m = df_weights['관리비'].reindex(self.hospital_types).fillna(0)
                        w_z = df_weights['재료비'].reindex(self.hospital_types).fillna(0)
                        
                        results[name] = (w_l * labor_rates[i_t] + 
                                       w_m * raw_rates[m_t] + 
                                       w_z * raw_rates[z_t])
            
            df = pd.DataFrame(results)
            df_final = pd.concat([df, pd.DataFrame({
                '평균': df.mean(axis=1), '최대': df.max(axis=1), 
                '최소': df.min(axis=1), '중위수': df.median(axis=1)
            })], axis=1)
            self._cache[target_year] = df_final # Store in cache
            return df_final
        except Exception as e:
            print(f"MEI 산출 중 오류 ({target_year}): {e}")
            return None

class SgrCalculator:
    def __init__(self, data, hospital_types, group_mapping=None):
        self.data = data
        self.hospital_types = hospital_types
        self.group_mapping = group_mapping or {}
        self._comp_cache = {} # [OPTIMIZATION] Memoization
        self._s1_cache = {}
        self._s2_cache = {}

    def _safe_get(self, df, year, col=None):
        try:
            y = max(df.index.min(), min(year, df.index.max()))
            if col: return df.loc[y, col]
            return df.loc[y]
        except: 
            return 1.0

    def _calc_sgr_components(self, year):
        if year in self._comp_cache:
            return self._comp_cache[year]
            
        try:
            gdp, pop, law, reval = self.data['df_gdp'], self.data['df_pop'], self.data['df_sgr_law'], self.data['df_sgr_reval']
            exp = self.data['df_expenditure']
            
            # Global (SGR uses per-capita GDP)
            denom = self._safe_get(gdp, year-1, '실질GDP') / self._safe_get(gdp, year-1, '영안인구')
            num = self._safe_get(gdp, year, '실질GDP') / self._safe_get(gdp, year, '영안인구')
            g_s1 = num / denom
            p_s1 = self._safe_get(pop, year, '건보대상자수') / self._safe_get(pop, year-1, '건보대상자수')
            g_s2 = 1 + (g_s1 - 1) * 0.8
            p_col_s2 = '건보_고령화반영후(대상자수)' if '건보_고령화반영후(대상자수)' in pop.columns else '건보대상자수'
            p_s2 = self._safe_get(pop, year, p_col_s2) / self._safe_get(pop, year-1, '건보대상자수')
            
            # Type Specific
            l = self._safe_get(law, year)
            r = self._safe_get(reval, year) / self._safe_get(reval, year-1)
            
            # Add Groups
            w_year = year - 2
            if w_year in exp.index:
                ae_w = exp.loc[w_year]
                for g, members in self.group_mapping.items():
                    valid = [m for m in members if m in l.index and m in ae_w.index]
                    if valid:
                        weights = ae_w[valid] / ae_w[valid].sum()
                        l[g] = (l[valid] * weights).sum()
                        r[g] = (r[valid] * weights).sum()
            
            res = {'g_s1': g_s1, 'p_s1': p_s1, 'g_s2': g_s2, 'p_s2': p_s2, 'l': l, 'r': r}
            self._comp_cache[year] = res # Store in cache
            return res
        except: return None

    def calc_sgr_index(self, components, model='S1'):
        if not components: return pd.Series(1.0, index=self.hospital_types + list(self.group_mapping.keys()))
        g = components['g_s1'] if model == 'S1' else components['g_s2']
        p = components['p_s1'] if model == 'S1' else components['p_s2']
        return g * p * components['l'] * components['r']

    def calc_paf_s1(self, target_year):
        if target_year in self._s1_cache:
             return self._s1_cache[target_year]
             
        ae_actual = self.data['df_expenditure']
        types_all = self.hospital_types + list(self.group_mapping.keys())
        
        # 1. Individual components
        y_recent = target_year - 2
        c_recent = self._calc_sgr_components(y_recent)
        idx_recent = self.calc_sgr_index(c_recent, model='S1')
        
        ae_22 = self._safe_get(ae_actual, y_recent-1)
        ae_23 = self._safe_get(ae_actual, y_recent)
        
        # 2. Accumulation Loop (Vectorized)
        y_start, y_end = target_year-11, target_year-2
        sums_tge = pd.Series(0.0, index=types_all)
        sums_ae = pd.Series(0.0, index=types_all)
        
        for y in range(y_start, y_end + 1):
            c = self._calc_sgr_components(y)
            idx = self.calc_sgr_index(c, model='S1')
            a_prev = self._safe_get(ae_actual, y-1)
            a_curr = self._safe_get(ae_actual, y)
            
            # Apply to individuals
            sums_tge[self.hospital_types] += a_prev[self.hospital_types] * idx[self.hospital_types]
            sums_ae[self.hospital_types] += a_curr[self.hospital_types]
            
            # Apply to groups (Aggregate sums)
            for g, members in self.group_mapping.items():
                sums_tge[g] += (a_prev[members] * idx[members]).sum()
                sums_ae[g] += a_curr[members].sum()

        # 3. Denominators
        c_24 = self._calc_sgr_components(target_year-1)
        idx_24 = self.calc_sgr_index(c_24, model='S1')
        denoms = pd.Series(0.0, index=types_all)
        
        denoms[self.hospital_types] = ae_23[self.hospital_types] * (1 + idx_24[self.hospital_types])
        for g, members in self.group_mapping.items():
             denoms[g] = (ae_23[members] * (1 + idx_24[members])).sum()

        # 4. Final UAF
        gap_short = pd.Series(0.0, index=types_all)
        tge_recent = ae_22 * idx_recent
        gap_short[self.hospital_types] = (tge_recent[self.hospital_types] - ae_23[self.hospital_types]) / ae_23[self.hospital_types]
        
        for g, members in self.group_mapping.items():
            agg_tge = (ae_22[members] * idx_recent[members]).sum()
            agg_ae = ae_23[members].sum()
            gap_short[g] = (agg_tge - agg_ae) / agg_ae

        gap_accum = (sums_tge - sums_ae) / denoms
        uaf = gap_short * 0.75 + gap_accum * 0.33
        self._s1_cache[target_year] = uaf
        return uaf

    def calc_paf_s2(self, target_year):
        if target_year in self._s2_cache:
             return self._s2_cache[target_year]
             
        # S2 usually follows similar aggregate logic for groups
        ae_actual = self.data['df_expenditure']
        types_all = self.hospital_types + list(self.group_mapping.keys())
        group_pafs = pd.Series(0.0, index=types_all)
        
        for lag, weight in [(2, 0.5), (3, 0.3), (4, 0.2)]:
            y = target_year - lag
            c = self._calc_sgr_components(y)
            idx = self.calc_sgr_index(c, model='S2')
            
            ae_prev = self._safe_get(ae_actual, y-1)
            ae_curr = self._safe_get(ae_actual, y)
            
            term_gaps = pd.Series(0.0, index=types_all)
            tge = ae_prev[self.hospital_types] * idx[self.hospital_types]
            term_gaps[self.hospital_types] = (tge - ae_curr[self.hospital_types]) / ae_curr[self.hospital_types]
                
            for g, members in self.group_mapping.items():
                agg_t = (ae_prev[members] * idx[members]).sum()
                agg_a = ae_curr[members].sum()
                term_gaps[g] = (agg_t - agg_a) / agg_a
                
            group_pafs += term_gaps * weight
            
        self._s2_cache[target_year] = group_pafs
        return group_pafs

class CalculationEngine:
    def __init__(self, baseline_data, overrides=None):
        self.HOSPITAL_TYPES = ['상급종합', '종합병원', '병원', '요양병원', '의원', '치과병원', '치과의원', '한방병원', '한의원', '약국']
        self.GROUP_MAPPING = {
            '병원(계)': ['상급종합', '종합병원', '병원', '요양병원'], '의원(계)': ['의원'],
            '치과(계)': ['치과병원', '치과의원'], '한방(계)': ['한방병원', '한의원'], '약국(계)': ['약국']
        }
        self.data = self._apply_overrides(baseline_data, overrides)
        self.mei_calc = MeiCalculator(self.data, self.HOSPITAL_TYPES)
        self.sgr_calc = SgrCalculator(self.data, self.HOSPITAL_TYPES, self.GROUP_MAPPING)
        
        # [Optimized] Pre-calculate expenditure weights for speed
        self._precalc_weights = {}
        for y in range(2010, 2030):
            try:
                exp = self.data['df_expenditure'].loc[y].reindex(self.HOSPITAL_TYPES).fillna(0)
                tot = exp.sum()
                if tot > 0: self._precalc_weights[y] = exp / tot
            except: pass

    def _apply_overrides(self, baseline, overrides):
        if not overrides: return baseline
        new_data = {k: v.copy() for k, v in baseline.items()}
        
        # Generic override injection for any year present in the overrides keys
        # Format expected: FIELD_YEAR (e.g., I1_2025, GDP_2024, etc.)
        for key, value in overrides.items():
            try:
                if '_' not in key: continue
                
                # Handling WEIGHT overrides (WEIGHT_{TYPE}_{COL})
                if key.startswith('WEIGHT_'):
                    parts = key.split('_')
                    if len(parts) == 3:
                        _, htype, col = parts
                        if htype in self.HOSPITAL_TYPES and col in ['인건비', '관리비', '재료비']:
                            new_data['df_weights'].loc[htype, col] = float(value)
                    continue

                field, year = key.rsplit('_', 1)
                year = int(year)
                
                # MEI Inflation
                if field in ['I1', 'I2', 'I3', 'M1', 'M2', 'Z1', 'Z2']:
                    col_map = {'I':'인건비','M':'관리비','Z':'재료비'}
                    col_name = f"{col_map[field[0]]}_{field[1]}"
                    if year in new_data['df_raw_mei_inf'].index:
                        new_data['df_raw_mei_inf'].loc[year, col_name] = float(value)
                
                # Macro Factors
                elif field == 'GDP':
                    new_data['df_gdp'].loc[year, '실질GDP'] = float(value)
                elif field == 'POP':
                    new_data['df_gdp'].loc[year, '영안인구'] = float(value)
                elif field == 'NHI_POP':
                    new_data['df_pop'].loc[year, '건보대상자수'] = float(value)
                
                # Law & RV (Type-specific)
                elif field.startswith('LAW_'):
                    parts = field.split('_')
                    if len(parts) == 2:
                        htype = parts[1]
                        if year in new_data['df_sgr_law'].index:
                            new_data['df_sgr_law'].loc[year, htype] = float(value)
                
                elif field.startswith('RV_'):
                    parts = field.split('_')
                    if len(parts) == 2:
                        htype = parts[1]
                        if year in new_data['df_rel_value'].index:
                            new_data['df_rel_value'].loc[year, htype] = float(value)

                elif field == 'LAW':
                    new_data['df_sgr_law'].loc[year] = float(value)
                elif field == 'RV':
                    # Relative value usually affects T-2 for CF calculation
                    if (year-2) in new_data['df_rel_value'].index:
                        new_data['df_rel_value'].loc[year-2] = 1 + float(value)/100

                # Medical Expenditure (Direct overrides)
                elif field in self.HOSPITAL_TYPES:
                    if year in new_data['df_expenditure'].index:
                        new_data['df_expenditure'].loc[year, field] = float(value)
                
                # CF Revaluation
                elif field.startswith('CF_') and field[3:] in self.HOSPITAL_TYPES:
                    actual_field = field[3:]
                    if year in new_data['df_sgr_reval'].index:
                        new_data['df_sgr_reval'].loc[year, actual_field] = float(value)

            except Exception as e:
                print(f"Override error for {key}: {e}")
                continue
                
        return new_data

    def calc_group_average(self, df_values, target_year):
        weight_year = target_year - 2
        weights = self._precalc_weights.get(weight_year)
        if weights is None: return pd.Series(0.0, index=list(self.GROUP_MAPPING.keys())+['전체'])

        # Filter values (Should be indexed by Hospital Types)
        v = df_values.reindex(self.HOSPITAL_TYPES).fillna(0.0)
        results = {}
        for group, members in self.GROUP_MAPPING.items():
            gw = weights.loc[members]
            g_sum = gw.sum()
            if g_sum > 0:
                results[group] = (v.loc[members] * (gw / g_sum)).sum()
            else:
                results[group] = v.loc[members].mean()
        
        results['전체'] = (v * weights).sum()
        return pd.Series(results)

    def run_full_analysis(self, target_year=2025):
        # 2014년부터 2028년까지 분석 (사용자 요청: 2010-2013 제외)
        years = range(2014, 2029) 
        print(f"[INFO] Analysis Range (Calculated): {list(years)}")
        history = {
            'years': list(years), 
            'S1': {}, 'S2': {}, 'Link': {}, 'MEI': {}, 'GDP': {}, 
            'UAF_S1': {}, 'UAF_S2': {},
            'SGR_S1_INDEX': {}, 'SGR_S2_INDEX': {},
            'Target_S1': {}, 'Target_S2': {},
            'S1_Rescaled': {}, 'S2_Rescaled': {},
            'IndexMethod': {}
        }
        details = {'mei_raw': {}, 'sgr_factors': {}}
        
        # Additional bulk data for 2020-2026 comparison
        bulk_sgr = {
            'reval_rates': {}, # 환산지수 변화율 (Index)
            'reval_growth': {}, # 환산지수 변화율 (%)
            'pop_growth': {},  # 인구증가율 (s1, s2)
            'law_changes': {}, # 법과제도
            'gdp_growth': {},   # GDP 증가율 (s1, s2, total)
            'mei_growth': {},   # MEI 증가율 (%)
            'factor_growth': {}, # 생산요소 물가 증가율
            'scenario_adjustments': {}, # 16가지 시나리오별 조정률
            'ar_analysis': {}, # AR모형 시나리오 분석 (30개)
            'budget_analysis': {} # 연도별 추가소요재정 분석
        }

        # Helper: Calculate Grouped Results (Rate & Budget)
        def calc_grouped_results(rates, base_amounts):
            # rates: Series of rates (e.g. 2.5 for 2.5%)
            # base_amounts: DataFrame or Series of base expenditure (e.g. 2024 expenditure)
            
            res = {'rate': {}, 'budget': {}}
            
            # Simple simulation: Budget = Base * (Rate/100)
            # Or if Rate is index change: Budget = Base * (1 + Rate/100) - Base = Base * Rate/100
            # Let's assume 'rates' are Percentage Increases (e.g. 1.6%).
            # Additional Budget = Base * (Rate / 100)
            
            # Individual types
            total_budget = 0
            for ht in self.HOSPITAL_TYPES:
                r = rates.get(ht, 0.0)
                base = base_amounts.get(ht, 0.0)
                add_budget = base * (r / 100)
                res['rate'][ht] = round(r, 2)
                res['budget'][ht] = round(add_budget, 0)
                total_budget += add_budget
            
            # Groups
            for grp, members in self.GROUP_MAPPING.items():
                grp_budget = 0
                grp_base_sum = 0
                for m in members:
                    r = rates.get(m, 0.0)
                    base = base_amounts.get(m, 0.0)
                    grp_budget += base * (r / 100)
                    grp_base_sum += base
                
                res['budget'][grp] = round(grp_budget, 0)
                if grp_base_sum > 0:
                    res['rate'][grp] = round((grp_budget / grp_base_sum) * 100, 2)
                else:
                    res['rate'][grp] = 0.0
            
            res['budget']['전체'] = round(total_budget, 0)
            total_base = sum(base_amounts.get(ht, 0.0) for ht in self.HOSPITAL_TYPES)
            if total_base > 0:
                 res['rate']['전체'] = round((total_budget / total_base) * 100, 2)
            else:
                 res['rate']['전체'] = 0.0
            
            return res
        
        # Helper: Combined for UAF (Rates) - already includes group results in the Series
        def get_rate_combined(rates):
            # rates is a Series indexed by both individuals and groups
            # Increase precision to 3 decimal places as requested
            return (rates * 100).round(3).to_dict()


        # Helper: Combined for generic values (MEI, etc) - still using AE weights for averages
        def get_combined(values, year):
            indiv = ((values[self.HOSPITAL_TYPES] - 1) * 100).round(2)
            groups = ((self.calc_group_average(values[self.HOSPITAL_TYPES], year) - 1) * 100).round(2)
            return {**indiv.to_dict(), **groups.to_dict()}

        # Helper: Indices (1.xxxx)
        def get_combined_index(values, year):
            indiv = values[self.HOSPITAL_TYPES].round(4)
            groups = self.calc_group_average(values[self.HOSPITAL_TYPES], year).round(4)
            return {**indiv.to_dict(), **groups.to_dict()}

        # Helper: Absolute (Target expenditures) - Sum for groups
        def get_combined_absolute(values, year):
            indiv = values[self.HOSPITAL_TYPES]
            results = {}
            for g, members in self.GROUP_MAPPING.items():
                results[g] = values[members].sum() if all(m in values.index for m in members) else 0
            results['전체'] = values[self.HOSPITAL_TYPES].sum()
            return {**indiv.round(0).to_dict(), **results}

        for y in years:
            # Initialize for safety (prevent JS error on frontend if calc fails)
            # 1. Initialize for the year y
            for k in ['S1', 'S2', 'Link', 'MEI', 'GDP', 'UAF_S1', 'UAF_S2', 'SGR_S1_INDEX', 'SGR_S2_INDEX', 'Target_S1', 'Target_S2', 'S1_Rescaled', 'S2_Rescaled']:
                history[k][y] = {}

            # 2. Base MEI Calculation (Current Year y gets T-2 data)
            df_mei = None
            mei_avg = None
            try:
                df_mei = self.mei_calc.calc_mei_index_by_year(y)
                if df_mei is not None:
                    details['mei_raw'][y] = df_mei.round(4).to_dict()
                    mei_avg = df_mei['평균']
            except Exception as e:
                print(f"[WARN] MEI calc failed for {y}: {e}")

            # 3. SGR Standard (S1) and Improved (S2) - FULL REVERT
            uaf_s1 = None
            uaf_s2 = None
            try:
                uaf_s1 = self.sgr_calc.calc_paf_s1(y)
                uaf_s2 = self.sgr_calc.calc_paf_s2(y)
                
                if mei_avg is not None:
                    if uaf_s1 is not None:
                        history['UAF_S1'][y] = get_rate_combined(uaf_s1)
                        cf_s1_idx = mei_avg * (1 + uaf_s1[self.HOSPITAL_TYPES])
                        history['S1'][y] = get_combined(cf_s1_idx, y)
                    
                    if uaf_s2 is not None:
                        history['UAF_S2'][y] = get_rate_combined(uaf_s2)
                        cf_s2_idx = mei_avg * (1 + uaf_s2[self.HOSPITAL_TYPES])
                        history['S2'][y] = get_combined(cf_s2_idx, y) # NO RV Deduction for SGR as requested
            except Exception as e:
                print(f"[WARN] SGR S1/S2 calc failed for {y}: {e}")

            # [NEW] Rescaling Logic (Rescale S1/S2 to match Contract Overall Rate)
            try:
                # Get Contract Overall Rate for Year y
                contract_rate = None
                df_contract = self.data.get('df_contract', pd.DataFrame())
                if not df_contract.empty and y in df_contract.index and '인상율_전체' in df_contract.columns:
                     val = df_contract.loc[y, '인상율_전체']
                     if pd.notna(val):
                         contract_rate = float(val)

                if contract_rate is not None and history['S1'][y] and '전체' in history['S1'][y]:
                    # Calculation of Current Overall Rates
                    # history['S1'][y] contains rounded percentages. For precision, let's recalculate unrounded overall if possible.
                    # Or just use the '전체' value from history which is the weighted average.
                    
                    # 1. S1 Rescaling
                    current_s1_avg = history['S1'][y]['전체'] # Percentage value (e.g. 2.05)
                    if abs(current_s1_avg) > 0.0001:
                        scale_factor_s1 = contract_rate / current_s1_avg
                        # Apply to all
                        rescaled_s1 = {k: v * scale_factor_s1 for k, v in history['S1'][y].items()}
                        # Recalculate '전체' just to be sure (it should be contract_rate)
                        rescaled_s1['전체'] = contract_rate
                        history['S1_Rescaled'][y] = {k: round(v, 2) for k, v in rescaled_s1.items()}
                    else:
                         history['S1_Rescaled'][y] = history['S1'][y] # No scaling if 0

                    # 2. S2 Rescaling
                    current_s2_avg = history['S2'][y]['전체']
                    if abs(current_s2_avg) > 0.0001:
                        scale_factor_s2 = contract_rate / current_s2_avg
                        rescaled_s2 = {k: v * scale_factor_s2 for k, v in history['S2'][y].items()}
                        rescaled_s2['전체'] = contract_rate
                        history['S2_Rescaled'][y] = {k: round(v, 2) for k, v in rescaled_s2.items()}
                    else:
                         history['S2_Rescaled'][y] = history['S2'][y]
                else:
                    # Fallback if no contract data or SGR calc failed
                    history['S1_Rescaled'][y] = history['S1'][y] if history['S1'][y] else {}
                    history['S2_Rescaled'][y] = history['S2'][y] if history['S2'][y] else {}

            except Exception as e_rescale:
                print(f"[WARN] Rescaling logic failed for {y}: {e_rescale}")
                history['S1_Rescaled'][y] = history['S1'][y] if history['S1'][y] else {}
                history['S2_Rescaled'][y] = history['S2'][y] if history['S2'][y] else {}

            # 4. Macro Indicator Models (GDP, MEI, Link) - Apply Simple Subtraction
            # [USER REQUEST] 2025 adjustment = (Indicator Index - RV Index) * 100
            # 복잡한 연산 없이 이미 산출된 지수 간의 차이만 가져와서 %로 표시 (매우 단순화)
            calc_year = y - 2
            try:
                # Get RV Index (Default: 1.0)
                rv_file = self.data['df_rel_value']
                rv_idx = pd.Series(1.0, index=self.HOSPITAL_TYPES)
                if calc_year in rv_file.index:
                    rv_idx = rv_file.loc[calc_year].reindex(self.HOSPITAL_TYPES).fillna(1.0)

                # 1. GDP Model
                gdp_file = self.data['df_gdp']
                g_idx = self.sgr_calc._safe_get(gdp_file, calc_year, '실질GDP') / self.sgr_calc._safe_get(gdp_file, calc_year-1, '실질GDP')
                
                # GDP Adjustment (%) = (GDP Index - RV Index) * 100
                indiv_gdp = ((g_idx - rv_idx) * 100).round(2)
                group_gdp = ((g_idx - self.calc_group_average(rv_idx, y)) * 100).round(2)
                history['GDP'][y] = {**indiv_gdp.to_dict(), **group_gdp.to_dict()}

                if mei_avg is not None:
                    # 2. MEI Model (Section 11)
                    # [USER REQUEST] mei_평균 index (1.0389 등) - rv_index (1.0 등)
                    m_idx = mei_avg 
                    indiv_mei = ((m_idx - rv_idx) * 100).round(2)
                    group_mei = ((self.calc_group_average(m_idx, y) - self.calc_group_average(rv_idx, y)) * 100).round(2)
                    history['MEI'][y] = {**indiv_mei.to_dict(), **group_mei.to_dict()}

                    # 3. Macro Link Model
                    # Link logic: G + 1/3*(M-G)
                    link_idx = m_idx.map(lambda m: g_idx + (1/3*(m-g_idx)) if m > g_idx else g_idx)
                    indiv_link = ((link_idx - rv_idx) * 100).round(2)
                    group_link = ((self.calc_group_average(link_idx, y) - self.calc_group_average(rv_idx, y)) * 100).round(2)
                    history['Link'][y] = {**indiv_link.to_dict(), **group_link.to_dict()}
            except Exception as e:
                print(f"[WARN] Macro models correction failed for {y}: {e}")

            # 5. SGR Index & Components for Current Year
            comp_y = None
            try:
                comp_y = self.sgr_calc._calc_sgr_components(y)
                if comp_y:
                    idx_s1 = self.sgr_calc.calc_sgr_index(comp_y, model='S1')
                    idx_s2 = self.sgr_calc.calc_sgr_index(comp_y, model='S2')
                    
                    history['SGR_S1_INDEX'][y] = get_combined_index(idx_s1, y)
                    history['SGR_S2_INDEX'][y] = get_combined_index(idx_s2, y)
                    
                    if (y-1) in self.data['df_expenditure'].index:
                        ae_prev = self.data['df_expenditure'].loc[y-1]
                        target_s1 = ae_prev * idx_s1
                        target_s2 = ae_prev * idx_s2
                        history['Target_S1'][y] = get_combined_absolute(target_s1, y)
                        history['Target_S2'][y] = get_combined_absolute(target_s2, y)

                    # Save Factors
                    factors_dict = {}
                    for k, v in comp_y.items():
                        if k in ['l', 'r'] and isinstance(v, (pd.Series, pd.DataFrame)):
                            pre_calc_val = None
                            for key_check in ['전체', '계', '평균']:
                                if key_check in v.index:
                                    pre_calc_val = float(v[key_check])
                                    break
                            
                            if pre_calc_val is not None:
                                factors_dict[k] = pre_calc_val
                            else:
                                avg_series = self.calc_group_average(v, y)
                                factors_dict[k] = float(avg_series['전체'])
                        elif hasattr(v, 'iloc'):
                            factors_dict[k] = float(v.iloc[0])
                        else:
                            factors_dict[k] = float(v)
                    details['sgr_factors'][y] = factors_dict
            except Exception as e:
                print(f"[WARN] SGR Components/Index calc failed for {y}: {e}")

            # 6. Index Method (Section 14)
            # [USER REQUEST] Index = MEI Scenario - (Revenue/Institutions Growth)
            try:
                calc_y = y - 2
                prev_y = y - 3
                df_exp = self.data['df_expenditure']
                df_num = self.data['df_num']
                
                if calc_y in df_exp.index and prev_y in df_exp.index and calc_y in df_num.index and prev_y in df_num.index:
                    # Calculate Individual Revenue Growth (%)
                    rev_per_inst_curr = df_exp.loc[calc_y, self.HOSPITAL_TYPES] / df_num.loc[calc_y, self.HOSPITAL_TYPES]
                    rev_per_inst_prev = df_exp.loc[prev_y, self.HOSPITAL_TYPES] / df_num.loc[prev_y, self.HOSPITAL_TYPES]
                    rev_growth = ((rev_per_inst_curr / rev_per_inst_prev) - 1) * 100
                    
                    # Store Revenue Growth (Indiv + Group)
                    indiv_rev = rev_growth.round(2).to_dict()
                    group_rev = self.calc_group_average(rev_growth / 100 + 1, y) # Weighting growth indices
                    group_rev_pct = ((group_rev - 1) * 100).round(2).to_dict()
                    
                    history['IndexMethod'][y] = {
                        'rev_growth': {**indiv_rev, **group_rev_pct},
                        'scenarios': {}
                    }
                    
                    if df_mei is not None:
                        for sn in df_mei.columns:
                            mei_growth_pct = (df_mei[sn] - 1) * 100
                            adj_rate = mei_growth_pct - rev_growth
                            
                            indiv_adj = adj_rate.round(2).to_dict()
                            group_adj_idx = self.calc_group_average(adj_rate / 100 + 1, y)
                            group_adj_pct = ((group_adj_idx - 1) * 100).round(2).to_dict()
                            
                            history['IndexMethod'][y]['scenarios'][sn] = {**indiv_adj, **group_adj_pct}
            except Exception as e:
                print(f"[WARN] Index Method calc failed for {y}: {e}")

            # 5. Bulk Data (Monitoring) - Range 2014-2028
            try:
                if 2014 <= y <= 2028:
                    
                    # Reval (상대가치 변화지수 원시값) 및 증가율
                    rv_raw = self.data['df_rel_value']
                    if y in rv_raw.index:
                        rv_vals = rv_raw.loc[y].reindex(self.HOSPITAL_TYPES).fillna(1.0)
                        bulk_sgr['reval_rates'][y] = get_combined_index(rv_vals, y)
                        bulk_sgr['reval_growth'][y] = get_combined(rv_vals, y)
                    
                    # MEI (해당 조정 연도 y에서 계산된 y-2년 MEI 평균 증가율 활용)
                    if mei_avg is not None:
                        bulk_sgr['mei_growth'][y-2] = get_combined(mei_avg, y)
                    
                    # Pop (건보대상자수 증가율)
                    pop = self.data['df_pop']
                    if y in pop.index and y-1 in pop.index:
                        p_cur = pop.loc[y, '건보대상자수']
                        p_prev = pop.loc[y-1, '건보대상자수']
                        s1_growth = (p_cur / p_prev - 1) * 100
                        
                        p_col_s2 = '건보_고령화반영후(대상자수)' if '건보_고령화반영후(대상자수)' in pop.columns else '건보대상자수'
                        s2_index_val = pop.loc[y, p_col_s2] / p_prev
                        s2_growth = (s2_index_val - 1) * 100
                        
                        bulk_sgr['pop_growth'][y] = {
                            's1': round(s1_growth, 3), 
                            's2': round(s2_growth, 3),
                            's2_index': round(s2_index_val, 4)
                        }
                    
                    # Law (법과제도 변화지수)
                    law = self.data['df_sgr_law']
                    if y in law.index:
                        indiv_law_idx = law.loc[y]
                        group_law_idx = self.calc_group_average(indiv_law_idx, y)
                        combined_law_idx = {**group_law_idx.to_dict(), **indiv_law_idx.to_dict()}
                        for key_check in ['전체', '계', '평균']:
                            if key_check in indiv_law_idx.index:
                                combined_law_idx['전체'] = indiv_law_idx[key_check]
                                break
                        bulk_sgr['law_changes'][y] = {k: round(v, 4) for k, v in combined_law_idx.items()}
                    
                    # GDP (1인당 실질 GDP 증가율 vs 실질 GDP 증가율 총액)
                    gdp = self.data['df_gdp']
                    if y in gdp.index and y-1 in gdp.index:
                        # 1인당 GDP 증가율 (SGR 모델용)
                        g1 = gdp.loc[y, '실질GDP'] / gdp.loc[y, '영안인구']
                        g0 = gdp.loc[y-1, '실질GDP'] / gdp.loc[y-1, '영안인구']
                        g_per_growth = (g1 / g0 - 1) * 100
                        
                        # GDP 총액 증가율 (거시 GDP 모형 및 모니터링용)
                        g_total_idx = gdp.loc[y, '실질GDP'] / gdp.loc[y-1, '실질GDP']
                        gdp_total_growth = (g_total_idx - 1) * 100
                        
                        bulk_sgr['gdp_growth'][y] = {
                            's1': round(g_per_growth, 3), 
                            's2': round(g_per_growth * 0.8, 3),
                            'total': round(gdp_total_growth, 3) 
                        }

                    # Factor Price (생산요소 물가 증가율)
                    inf = self.data['df_raw_mei_inf']
                    if y in inf.index:
                        f_growth = {}
                        # Labor (3yr Geometric Mean Growth)
                        if y-3 in inf.index:
                            for i in [1, 2, 3]:
                                col = f'인건비_{i}'
                                if col in inf.columns:
                                    rate = (inf.loc[y, col] / inf.loc[y-3, col])**(1/3) - 1
                                    f_growth[col] = round(rate * 100, 3)
                        # Admin & Material (Prev Year Growth)
                        if y-1 in inf.index:
                            for prefix in ['관리비', '재료비']:
                                for i in [1, 2]:
                                    col = f'{prefix}_{i}'
                                    if col in inf.columns:
                                        rate = (inf.loc[y, col] / inf.loc[y-1, col]) - 1
                                        f_growth[col] = round(rate * 100, 3)
                        bulk_sgr['factor_growth'][y] = f_growth

                    # [OPTIMIZED] Scenario Adjustments - Vectorized
                    if df_mei is not None and uaf_s1 is not None and uaf_s2 is not None:
                        summary_cols = ['평균', '최대', '최소', '중위수']
                        other_cols = [c for c in df_mei.columns if c not in summary_cols]
                        ordered_cols = summary_cols + other_cols
                        
                        bulk_sgr['scenario_adjustments'][y] = {}
                        
                        # Calculate all scenarios for S1 and S2 at once
                        w_year = y - 2
                        weights = self._precalc_weights.get(w_year)
                        
                        if weights is not None:
                            # Pre-calculate base adjustment series
                            # S1: CF = MEI * (1+UAF_S1)
                            # S2: CF = MEI * (1+UAF_S2)
                            h_types = self.HOSPITAL_TYPES
                            s1_adj = 1 + uaf_s1[h_types]
                            s2_adj = 1 + uaf_s2[h_types]
                            
                            # Matrix operations for all scenarios
                            mei_matrix = df_mei[ordered_cols].reindex(h_types).fillna(1.0)
                            cf_s1_matrix = mei_matrix.mul(s1_adj, axis=0)
                            cf_s2_matrix = mei_matrix.mul(s2_adj, axis=0)
                            
                            # Percentage terms: (CF - 1)*100
                            cf_s1_pct = ((cf_s1_matrix - 1) * 100).round(2)
                            cf_s2_pct = ((cf_s2_matrix - 1) * 100).round(2)
                            
                            # Group averages for all scenarios
                            for sn in ordered_cols:
                                res_s1 = cf_s1_pct[sn].to_dict()
                                res_s2 = cf_s2_pct[sn].to_dict()
                                
                                # Add group averages
                                for g, members in self.GROUP_MAPPING.items():
                                    gw = weights.loc[members]
                                    g_sum = gw.sum()
                                    if g_sum > 0:
                                        res_s1[g] = round((cf_s1_pct.loc[members, sn] * (gw / g_sum)).sum(), 2)
                                        res_s2[g] = round((cf_s2_pct.loc[members, sn] * (gw / g_sum)).sum(), 2)
                                    else:
                                        res_s1[g] = round(cf_s1_pct.loc[members, sn].mean(), 2)
                                        res_s2[g] = round(cf_s2_pct.loc[members, sn].mean(), 2)
                                
                                # Total average
                                res_s1['전체'] = round((cf_s1_pct.loc[h_types, sn] * weights).sum(), 2)
                                res_s2['전체'] = round((cf_s2_pct.loc[h_types, sn] * weights).sum(), 2)
                                
                                bulk_sgr['scenario_adjustments'][y][sn] = {'S1': res_s1, 'S2': res_s2}
                        else:
                            # Fallback to slower method if no weights
                            for sn in ordered_cols:
                                if sn in df_mei.columns:
                                    bulk_sgr['scenario_adjustments'][y][sn] = {
                                        'S1': get_combined(df_mei[sn] * (1 + uaf_s1), y),
                                        'S2': get_combined(df_mei[sn] * (1 + uaf_s2), y)
                                    }

                        # --- [Optimized] AR Model Scenario Analysis (2020-2028) ---
                        if 2020 <= y <= 2028:
                            ar_results_all = {}
                            # Load weights once per year
                            try:
                                exp_w = self.data['df_expenditure'].loc[y-2].reindex(self.HOSPITAL_TYPES).fillna(0)
                                total_exp = exp_w.sum()
                                if total_exp > 0:
                                    normalized_w = exp_w / total_exp
                                else:
                                    normalized_w = pd.Series(1.0/len(self.HOSPITAL_TYPES), index=self.HOSPITAL_TYPES)
                            except:
                                normalized_w = pd.Series(1.0/len(self.HOSPITAL_TYPES), index=self.HOSPITAL_TYPES)

                            base_rate_keys = ['GDP', 'MEI', 'Link']
                            mei_scenario_keys = ['I1M2Z2', '평균']
                            r_values = [1.0, 0.75, 0.5, 0.25, 0.15, 0.0]

                            for m_key in ['S1', 'S2']:
                                ar_results = []
                                for br_key in base_rate_keys:
                                    if y in history[br_key] and history[br_key][y]:
                                        # Pre-convert base rates to Series
                                        base_rates_s = pd.Series({t: history[br_key][y].get(t, 0) for t in self.HOSPITAL_TYPES})
                                        
                                        for mei_sn in mei_scenario_keys:
                                            if mei_sn in bulk_sgr['scenario_adjustments'][y]:
                                                cf_s_dict = bulk_sgr['scenario_adjustments'][y][mei_sn][m_key]
                                                cf_s_s = pd.Series({t: cf_s_dict.get(t, 0) for t in self.HOSPITAL_TYPES})
                                                
                                                # Weighted average CF_S & adjustment vector
                                                avg_cf_s = (cf_s_s * normalized_w).sum()
                                                cf_adj = cf_s_s - avg_cf_s
                                                
                                                for r in r_values:
                                                    # Vectorized result calculation
                                                    final_rates_s = base_rates_s + r * cf_adj
                                                    
                                                    # Result composition
                                                    res_combined = {t: round(v, 2) for t, v in final_rates_s.to_dict().items()}
                                                    
                                                    for group, members in self.GROUP_MAPPING.items():
                                                        m_w = normalized_w.loc[members]
                                                        sum_w = m_w.sum()
                                                        if sum_w > 0:
                                                            res_combined[group] = round((final_rates_s.loc[members] * (m_w/sum_w)).sum(), 2)
                                                    
                                                    res_combined['전체'] = round((final_rates_s * normalized_w).sum(), 2)
                                                    
                                                    ar_results.append({
                                                        'base_rate': br_key,
                                                        'mei_scenario': mei_sn,
                                                        'r': r,
                                                        'rates': res_combined
                                                    })
                                ar_results_all[m_key] = ar_results
                            bulk_sgr['ar_analysis'][y] = ar_results_all

                            history_constrained = {}
                            df_contract = self.data['df_contract']
                            base_y = y - 2
                            rate_py = pd.Series(0.75, index=self.HOSPITAL_TYPES)
                            if 'df_rate_py' in self.data and base_y in self.data['df_rate_py'].index:
                                rate_py = self.data['df_rate_py'].loc[base_y].reindex(self.HOSPITAL_TYPES).fillna(0.75)
                            
                            base_amt = pd.Series(0.0, index=self.HOSPITAL_TYPES)
                            if base_y in self.data['df_expenditure'].index:
                                raw_exp = self.data['df_expenditure'].loc[base_y].reindex(self.HOSPITAL_TYPES).fillna(0)
                                base_amt = raw_exp * rate_py

                            budget_data = {'Macro': {}, 'S1': {}, 'S2': {}} 

                            # --- [NEW] 1. Macro Baselines ---
                            for macro_key in ['GDP', 'MEI', 'Link']:
                                if y in history[macro_key] and history[macro_key][y]:
                                    m_rates = pd.Series(history[macro_key][y]) # Percents
                                    budget_data['Macro'][macro_key] = calc_grouped_results(m_rates, base_amt)

                            # --- [NEW] 2. SGR AR Baselines (Target r=0.15, MEI=Average) --- 
                            target_r = 0.15
                            target_mei = '평균'
                            
                            for m_key in ['S1', 'S2']:
                                ar_list = ar_results_all.get(m_key, [])
                                
                                # Find AR1 (GDP), AR2 (MEI), AR3 (Link)
                                ar1 = next((x for x in ar_list if x['base_rate'] == 'GDP' and x['mei_scenario'] == target_mei and abs(x['r']-target_r)<0.001), None)
                                ar2 = next((x for x in ar_list if x['base_rate'] == 'MEI' and x['mei_scenario'] == target_mei and abs(x['r']-target_r)<0.001), None)
                                ar3 = next((x for x in ar_list if x['base_rate'] == 'Link' and x['mei_scenario'] == target_mei and abs(x['r']-target_r)<0.001), None)
                                
                                if ar1: budget_data[m_key]['AR1'] = calc_grouped_results(pd.Series(ar1['rates']), base_amt)
                                if ar2: budget_data[m_key]['AR2'] = calc_grouped_results(pd.Series(ar2['rates']), base_amt)
                                if ar3: budget_data[m_key]['AR3'] = calc_grouped_results(pd.Series(ar3['rates']), base_amt)
                                
                                if ar1 and ar2 and ar3:
                                    # AR Average
                                    r1 = pd.Series(ar1['rates'])
                                    r2 = pd.Series(ar2['rates'])
                                    r3 = pd.Series(ar3['rates'])
                                    avg_rates = (r1 + r2 + r3) / 3.0
                                    budget_data[m_key]['AR_Average'] = calc_grouped_results(avg_rates, base_amt)
                            
                            # --- [NEW] 3. Budget Constraints Analysis (5 Scenarios) ---
                            # Generalized for any year y
                            try:
                                # Initialize bulk storage for this year if not exists
                                if 'budget_constraints' not in bulk_sgr:
                                    bulk_sgr['budget_constraints'] = {}
                                
                                # Use y-1 as the base for lookback (previous year)
                                py = y - 1
                                
                                # Check if we have sufficient contract history
                                # We need at least up to y-5 for 5-year scenarios
                                has_history = (py in df_contract.index)
                                
                                if has_history:
                                    # Reference for Scaling: S2 AR Average (Total)
                                    ref_budget = 0
                                    ref_rate = 0
                                    
                                    if 'S2' in budget_data and 'AR_Average' in budget_data['S2']:
                                        ref_budget = budget_data['S2']['AR_Average']['budget'].get('전체', 0)
                                        ref_rate = budget_data['S2']['AR_Average']['rate'].get('전체', 0)
                                    elif 'S1' in budget_data and 'AR_Average' in budget_data['S1']:
                                        ref_budget = budget_data['S1']['AR_Average']['budget'].get('전체', 0)
                                        ref_rate = budget_data['S1']['AR_Average']['rate'].get('전체', 0)
                                    
                                    if ref_budget > 0 and ref_rate > 0:
                                        # Historical Data Fetching
                                        # Budget (Additional Finance)
                                        b_py = df_contract.loc[py, '추가소요재정_전체'] if py in df_contract.index else 0
                                        b_py_3 = df_contract.loc[py-3, '추가소요재정_전체'] if (py-3) in df_contract.index else 0
                                        b_py_4 = df_contract.loc[py-4, '추가소요재정_전체'] if (py-4) in df_contract.index else 0
                                        
                                        # Rate (Increase Rate)
                                        r_py = df_contract.loc[py, '인상율_전체'] if py in df_contract.index else 0
                                        r_py_1 = df_contract.loc[py-1, '인상율_전체'] if (py-1) in df_contract.index else 0
                                        r_py_2 = df_contract.loc[py-2, '인상율_전체'] if (py-2) in df_contract.index else 0
                                        r_py_3 = df_contract.loc[py-3, '인상율_전체'] if (py-3) in df_contract.index else 0
                                        r_py_4 = df_contract.loc[py-4, '인상율_전체'] if (py-4) in df_contract.index else 0

                                        # --- [Revised] Scenarios Definitions ---
                                        # Instead of one scale factor, we define the TARGET value for each scenario.
                                        targets = {}

                                        # S1.1: 5y Budget CAGR (py-4 -> py) applied to py
                                        if b_py_4 > 0 and b_py > 0:
                                            cagr_11 = (b_py / b_py_4)**(1/4) - 1
                                            target_b_11 = b_py * (1 + cagr_11)
                                            targets['S1_1'] = {'type': 'budget', 'value': target_b_11}
                                        else: 
                                            targets['S1_1'] = {'type': 'budget', 'value': b_py}

                                        # S1.2: 4y Budget CAGR (py-3 -> py) applied to py
                                        if b_py_3 > 0 and b_py > 0:
                                            cagr_12 = (b_py / b_py_3)**(1/3) - 1
                                            target_b_12 = b_py * (1 + cagr_12)
                                            targets['S1_2'] = {'type': 'budget', 'value': target_b_12}
                                        else:
                                            targets['S1_2'] = {'type': 'budget', 'value': b_py}

                                        # S2.1: 5y Rate Avg (py-4 to py)
                                        avg_r_21 = (r_py_4 + r_py_3 + r_py_2 + r_py_1 + r_py) / 5
                                        targets['S2_1'] = {'type': 'rate', 'value': avg_r_21}

                                        # S2.2: 3y Rate Avg (py-2 to py)
                                        avg_r_22 = (r_py_2 + r_py_1 + r_py) / 3
                                        targets['S2_2'] = {'type': 'rate', 'value': avg_r_22}

                                        # S2.3: Prev Year Rate (py)
                                        targets['S2_3'] = {'type': 'rate', 'value': r_py}

                                        # --- Apply Constraints to EACH Model individually ---
                                        # For each scenario, we create a full copy of budget_data structure, 
                                        # but every single model inside it is re-scaled to match the target.
                                        
                                        history_constrained_year = {}

                                        for s_key, t_info in targets.items():
                                            target_val = t_info['value']
                                            target_type = t_info['type']
                                            
                                            # Create a new structure for this scenario
                                            scenario_result = {}
                                            
                                            for cat, sub in budget_data.items(): # Macro, S1, S2
                                                scenario_result[cat] = {}
                                                for model_name, content in sub.items(): # GDP, AR1, AR_Average...
                                                    # content = {'rate': {type: val}, 'budget': {type: val}}
                                                    
                                                    # 1. Determine current total for this specific model
                                                    current_total_budget = content['budget'].get('전체', 0)
                                                    current_avg_rate = content['rate'].get('전체', 0)
                                                    
                                                    # 2. Calculate k (scaling factor) for THIS model
                                                    k = 1.0
                                                    if target_type == 'budget':
                                                        if current_total_budget > 0:
                                                            k = target_val / current_total_budget
                                                    elif target_type == 'rate':
                                                        if current_avg_rate > 0:
                                                            k = target_val / current_avg_rate
                                                    
                                                    # 3. Apply k to create new content
                                                    new_content = {'rate': {}, 'budget': {}}
                                                    
                                                    # Scaling Logic:
                                                    # If we scale Rate by k, Budget scales roughly by k (Budget ~ Rate * Base).
                                                    # If we scale Budget by k, Rate scales roughly by k.
                                                    # So valid for both: New_Value = Old_Value * k
                                                    
                                                    for ht, val in content['rate'].items():
                                                        new_content['rate'][ht] = round(val * k, 2)
                                                    
                                                    for ht, val in content['budget'].items():
                                                        new_content['budget'][ht] = round(val * k, 0)
                                                    
                                                    # Force the Total to be exactly target_val? 
                                                    # Rounding might cause slight drift, but k-scaling is the standard method.
                                                    # We will trust the scaled values.
                                                    
                                                    scenario_result[cat][model_name] = new_content
                                            
                                            history_constrained_year[s_key] = scenario_result

                                        # Store in bulk_sgr
                                        bulk_sgr['budget_constraints'][y] = history_constrained_year
                                        
                            except Exception as e:
                                print(f"Error calculating budget constraints for {y}: {str(e)}")
                                import traceback
                                traceback.print_exc()


                            bulk_sgr.setdefault('budget_analysis', {})[y] = budget_data
                                    
            except Exception as e:
                print(f"[WARN] Bulk SGR calc failed for {y}: {e}")

        return history, details, bulk_sgr

# ----------------------------------------------------------------------
# 2.5 AI Optimization Integrated via ai_optimizer.py


# ----------------------------------------------------------------------
# 3. Flask Server
# ----------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = get_secret('flask.secret_key', 'sgr_analytics_secret_safe_key') # 보안을 위한 시크릿 키

# 로그인 확인 데코레이터
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login')
def login():
    """로그인 페이지 (Firebase 설정을 동적으로 주입)"""
    firebase_config = {
        "apiKey": get_secret("firebase.apiKey"),
        "authDomain": get_secret("firebase.authDomain"),
        "projectId": get_secret("firebase.projectId"),
        "storageBucket": get_secret("firebase.storageBucket"),
        "messagingSenderId": get_secret("firebase.messagingSenderId"),
        "appId": get_secret("firebase.appId"),
        "measurementId": get_secret("firebase.measurementId")
    }
    return render_template('login.html', firebase_config=firebase_config)

@app.route('/logout')
def logout():
    """로그아웃"""
    session.pop('user', None)
    return redirect(url_for('login'))

@app.route('/set_session', methods=['POST'])
def set_session():
    """프론트엔드에서 파이어베이스 로그인 성공 시 세션 설정"""
    data = request.json
    token = data.get('token')
    email = data.get('email')
    
    # 이메일 화이트리스트 검종
    if email == 'fapitta1346@gmail.com':
        session['user'] = email
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Unauthorized email'}), 403

processor = DataProcessor('SGR_data.xlsx')

# 전역 캐시 변수 - 초기 로딩 시간 단축
_cached_analysis = None

def get_cached_analysis(force_reload=False):
    """캐시된 분석 결과를 반환하거나 새로 계산"""
    global _cached_analysis
    
    if force_reload:
        print("[INFO] 강제 데이터 새로고침 실행 중...")
        processor.reload_data()
        _cached_analysis = None
    
    if _cached_analysis is None:
        print("[INFO] 초기 분석 실행 중...")
        calc_engine = CalculationEngine(processor.raw_data)
        history, components, bulk_sgr = calc_engine.run_full_analysis(target_year=2025)
        
        # --- AI Integration Step (High Performance) ---
        try:
            if AI_MODULE_AVAILABLE:
                print("[Info] Running AI Optimization on App Load...")
                ai_engine = AIOptimizationEngine(data_frames=processor.raw_data)
                
                # Fetch S1 Model, MEI-Average results as baseline for year 2026
                # (Assuming bulk_sgr['scenario_adjustments'][2026]['평균']['S1'] exists)
                sgr_ref = {}
                try:
                    target_y = 2026
                    if 'scenario_adjustments' in bulk_sgr and target_y in bulk_sgr['scenario_adjustments']:
                        sgr_ref = bulk_sgr['scenario_adjustments'][target_y].get('평균', {}).get('S1', {})
                    
                    # If empty, fallback to simple history or defaults
                    if not sgr_ref:
                        sgr_ref = {'병원(계)': 1.96, '의원': 1.9, '치과(계)': 1.96, '한방(계)': 1.96, '약국': 2.8}
                except:
                    sgr_ref = {}

                ai_results = ai_engine.run_full_analysis(target_year=2026, sgr_results=sgr_ref)
                if ai_results:
                    bulk_sgr['ai_prediction'] = ai_results
                    print("[SUCCESS] AI Optimization analysis integrated with S1 Reference.")
        except Exception as e:
            print(f"[Error] AI Auto-run Failed: {e}")

        groups_list = calc_engine.HOSPITAL_TYPES + ['병원(계)', '의원(계)', '치과(계)', '한방(계)', '약국(계)', '전체']
        scenarios_list = list(components['mei_raw'][2025].keys()) if 2025 in components['mei_raw'] else ['평균', '최대', '최소', '중위수']
        
        _cached_analysis = {
            'history': history,
            'components': components,
            'bulk_sgr': bulk_sgr,
            'groups': groups_list,
            'scenarios': scenarios_list,
            'model_name_map': {
                'SGR_S1': 'S1', 'SGR_S2': 'S2', 'MACRO_GDP': 'GDP', 'MACRO_MEI': 'MEI', 'MACRO_LINK': 'Link'
            }
        }
        print("[SUCCESS] 초기 분석 완료!")
    
    return _cached_analysis

# Recursive check for NaN/Inf/Numpy types to prevent JSON errors
def sanitize_data(obj):
    if isinstance(obj, dict):
        return {str(k): sanitize_data(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_data(x) for x in obj]
    elif isinstance(obj, (bool, np.bool_)):
        return bool(obj)
    elif isinstance(obj, (int, np.integer)):
        return int(obj)
    elif isinstance(obj, (float, np.floating)):
        if pd.isna(obj) or np.isinf(obj): return None
        return float(obj)
    elif hasattr(obj, 'to_dict'):
        return sanitize_data(obj.to_dict())
    elif isinstance(obj, str):
        # [USER REQUEST] Clean HTML tags
        return re.sub(r'<[^>]+>', '', obj)
    return obj

@app.route('/')
def landing_redirect():
    """Redirect to main application (skipping landing page as requested)"""
    if 'user' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('index'))

@app.route('/ai')
@login_required
def ai_dashboard():
    """AI 최적화 대시보드 페이지"""
    return render_template('ai_dashboard.html')

@app.route('/app')
@login_required
def index():
    """Main application page - Runs analysis on first dashboard load"""
    selected_model = request.args.get('model', 'SGR_S2')
    tab = request.args.get('tab', 'dashboard')
    
    if tab == 'dashboard':
        # 대시보드 탭: 서버 측 캐시 활용 (사용량 절감)
        global _cached_analysis
        _cached_analysis = get_cached_analysis(force_reload=False)
        analysis_data = _cached_analysis.copy()
    else:
        # 입력 탭: 원시 데이터만 필요 (빠른 로딩)
        analysis_data = {
            'history': {'years': list(range(2014, 2029))},
            'components': {},
            'bulk_sgr': {},
            'groups': ['상급종합', '종합병원', '병원', '요양병원', '의원', '치과병원', '치과의원', '한방병원', '한의원', '약국', '병원(계)', '의원(계)', '치과(계)', '한방(계)', '약국(계)', '전체'],
            'scenarios': ['평균', '최대', '최소', '중위수'],
            'model_name_map': {
                'SGR_S1': 'S1', 'SGR_S2': 'S2', 'MACRO_GDP': 'GDP', 'MACRO_MEI': 'MEI', 'MACRO_LINK': 'Link'
            }
        }
    
    analysis_data['selected_model'] = selected_model
    analysis_data = sanitize_data(analysis_data)
    
    return render_template('index.html', analysis_data=analysis_data)

@app.route('/simulate', methods=['POST'])
@login_required
def simulate():
    overrides = request.json
    engine = CalculationEngine(processor.raw_data, overrides)
    history, components, bulk_sgr = engine.run_full_analysis()
    
    # Dynamic Scenario List
    scenarios_list = list(components['mei_raw'][2025].keys()) if 2025 in components['mei_raw'] else []
    if not scenarios_list:
        scenarios_list = ['평균', '최대', '최소', '중위수']

    response_data = {
        'success': True,
        'analysis_data': {
            'history': history,
            'components': components,
            'bulk_sgr': bulk_sgr,
            'groups': engine.HOSPITAL_TYPES + ['병원(계)', '의원(계)', '치과(계)', '한방(계)', '약국(계)', '전체'],
            'scenarios': scenarios_list,
            'model_name_map': {
                'SGR_S1': 'S1', 'SGR_S2': 'S2', 'MACRO_GDP': 'GDP', 'MACRO_MEI': 'MEI', 'MACRO_LINK': 'Link'
            }
        }
    }
    return jsonify(sanitize_data(response_data))

@app.route('/run_analysis', methods=['POST'])
@login_required
def run_analysis():
    """[NEW] 분석 실행 엔드포인트 - 사용자가 버튼을 클릭할 때만 실행"""
    try:
        print("[INFO] 사용자 요청으로 분석 실행 중...")
        global _cached_analysis
        
        # 전체 분석 실행
        engine = CalculationEngine(processor.raw_data)
        history, components, bulk_sgr = engine.run_full_analysis(target_year=2025)
        
        groups_list = engine.HOSPITAL_TYPES + ['병원(계)', '의원(계)', '치과(계)', '한방(계)', '약국(계)', '전체']
        scenarios_list = list(components['mei_raw'][2025].keys()) if 2025 in components['mei_raw'] else ['평균', '최대', '최소', '중위수']
        
        # 캐시 업데이트
        _cached_analysis = {
            'history': history,
            'components': components,
            'bulk_sgr': bulk_sgr,
            'groups': groups_list,
            'scenarios': scenarios_list,
            'model_name_map': {
                'SGR_S1': 'S1', 'SGR_S2': 'S2', 'MACRO_GDP': 'GDP', 'MACRO_MEI': 'MEI', 'MACRO_LINK': 'Link'
            }
        }
        
        print("[SUCCESS] 분석 완료 및 캐시 업데이트!")
        
        return jsonify({
            'success': True,
            'analysis_data': sanitize_data(_cached_analysis)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/sync_data', methods=['POST'])
@login_required
def sync_data():
    """사용자가 버튼을 눌러 구글 시트 데이터를 수동으로 갱신하는 엔드포인트"""
    try:
        print("[INFO] 구글 시트 데이터 수동 동기화 시작...")
        global _cached_analysis
        
        # 1. 프로세서에서 구글 시트 데이터 강제 새로고침
        processor.reload_data()
        
        # 2. 전역 분석 캐시 무효화 및 강제 재산출
        _cached_analysis = get_cached_analysis(force_reload=True)
        
        print("[SUCCESS] 수동 동기화 완료!")
        return jsonify({
            'success': True,
            'message': '구글 시트의 최신 데이터를 가져왔으며 분석 결과가 갱신되었습니다.'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/save_to_excel_file', methods=['POST'])
@login_required
def save_to_excel_file():
    data = request.json
    overrides = data.get('overrides', data)
    mode = data.get('mode', 'final')
    
    success, msg = processor.save_overrides_to_excel(overrides, mode=mode)
    if success:
        if mode == 'final':
            # Re-run full analysis with new base data
            global _cached_analysis
            _cached_analysis = None # invalidate cache
        return jsonify({'success': True, 'message': msg})
    else:
        return jsonify({'success': False, 'error': msg})

@app.route('/download_ar/<int:year>/<string:model_type>')
@login_required
def download_ar(year, model_type):
    """AR 모형 시나리오 분석 결과를 엑셀로 내보내기"""
    analysis = get_cached_analysis()
    ar_data_all = analysis['bulk_sgr']['ar_analysis'].get(year, {})
    
    # If it's the old list format (for backward compatibility during dev) or missing
    if isinstance(ar_data_all, list) and model_type == 'S1':
        ar_data = ar_data_all
    else:
        ar_data = ar_data_all.get(model_type, [])
    
    if not ar_data:
        return f"{model_type} 모델에 대한 데이터가 없습니다. (2020-2028년 범위 내에서만 지원됩니다.)", 404

    # 데이터프레임 생성
    rows = []
    for d in ar_data:
        row = {
            '거시지표(B)': d['base_rate'],
            'MEI(S)': d['mei_scenario'],
            '적용률(r)': d['r']
        }
        # Rates 추가
        for k, v in d['rates'].items():
            row[k] = v
        rows.append(row)
    
    df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'AR_Scenario_{year}')
        
        # 워크시트 스타일링 (가독성 향상)
        ws = writer.sheets[f'AR_Scenario_{year}']
        # 열 너비 조정
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"AR_Scenario_Analysis_{year}.xlsx")


@app.route('/download_budget/<int:year>')
def download_budget(year):
    """연구수가 및 추가소요재정 분석 결과를 엑셀로 내보내기"""
    analysis = get_cached_analysis()
    b_data = analysis['bulk_sgr'].get('budget_analysis', {}).get(year)
    
    if not b_data:
        return f"해당 연도({year}년)의 분석 데이터가 없습니다.", 404

    rows = []
    # 1. Macro baseline
    for m_key, data in b_data.get('Macro', {}).items():
        # Rate row
        r_row = {'모델': 'Macro 기초모형', '시나리오': m_key, '구분': '조정률(%)'}
        r_row.update(data['rate'])
        rows.append(r_row)
        # Budget row
        b_row = {'모델': 'Macro 기초모형', '시나리오': m_key, '구분': '소요재정(억)'}
        b_row.update(data['budget'])
        rows.append(b_row)

    # 2. S1 / S2
    for model in ['S1', 'S2']:
        m_label = '현행 SGR (S1)' if model == 'S1' else 'SGR 개선 (S2)'
        for s_key, data in b_data.get(model, {}).items():
            # Rate row
            r_row = {'모델': m_label, '시나리오': s_key, '구분': '조정률(%)'}
            r_row.update(data['rate'])
            rows.append(r_row)
            # Budget row
            b_row = {'모델': m_label, '시나리오': s_key, '구분': '소요재정(억)'}
            b_row.update(data['budget'])
            rows.append(b_row)

    df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = f'Budget_Analysis_{year}'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        ws = writer.sheets[sheet_name]
        # Column width adjustment
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'SGR_Budget_Analysis_{year}.xlsx'
    )


@app.route('/download_budget_constrained')
def download_budget_constrained():
    """추가소요재정 제약하의 분석 결과를 엑셀로 내보내기"""
    analysis = get_cached_analysis()
    c_data = analysis['bulk_sgr'].get('budget_constraints')
    
    if not c_data:
        return "추가소요재정 제약 분석 데이터가 없습니다.", 404

    rows = []
    scenario_names = {
        'S1_1': '시나리오 1.1 (5개년 재정증가율 반영)',
        'S1_2': '시나리오 1.2 (4개년 재정증가율 반영)',
        'S2_1': '시나리오 2.1 (5개년 평균인상율 반영)',
        'S2_2': '시나리오 2.2 (3개년 평균인상율 반영)',
        'S2_3': '시나리오 2.3 (직전연도 인상율 반영)'
    }

    for key, data in c_data.items():
        label = scenario_names.get(key, key)
        # Rate row
        r_row = {'시나리오': label, '구분': '조정률(%)'}
        r_row.update(data['rate'])
        rows.append(r_row)
        # Budget row
        b_row = {'시나리오': label, '구분': '소요재정(억)'}
        b_row.update(data['budget'])
        rows.append(b_row)

    df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Budget_Constrained_Analysis'
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 2

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='SGR_Budget_Constrained_Analysis_2025.xlsx'
    )
@app.route('/get_original_data')
def get_original_data():
    """원본 데이터를 JSON 형식으로 반환"""
    # Helper to safely convert values to JSON-compliant float or None
    def clean_val(x):
        try:
            if pd.isna(x): return None
            v = float(x)
            if np.isinf(v) or np.isnan(v): return None
            return v
        except:
            return None

    print("[API] get_original_data called")
    try:
        # [OPTIMIZATION] Avoid redundant disk reload. Data is already managed by processor.
        # processor.reload_data() 
        years = list(range(2010, 2029))  # 2010년부터 2028년까지 전체 기간으로 확장
        
        # MEI 물가지수 데이터
        mei_data = {}
        df_mei = processor.raw_data['df_raw_mei_inf']
        for field in ['인건비_1', '인건비_2', '인건비_3', '관리비_1', '관리비_2', '재료비_1', '재료비_2']:
            mei_data[field] = {}
            for year in years:
                if year in df_mei.index and field in df_mei.columns:
                    val = df_mei.loc[year, field]
                    mei_data[field][str(year)] = clean_val(val)
                else:
                    mei_data[field][str(year)] = None
        
        # 실제진료비 데이터
        medical_data = {}
        df_exp = processor.raw_data['df_expenditure']
        hospital_types = ['상급종합', '종합병원', '병원', '요양병원', '의원', '치과병원', '치과의원', '한방병원', '한의원', '약국']
        for htype in hospital_types:
            medical_data[htype] = {}
            for year in years:
                if year in df_exp.index and htype in df_exp.columns:
                    val = df_exp.loc[year, htype]
                    medical_data[htype][str(year)] = clean_val(val)
                else:
                    medical_data[htype][str(year)] = None
        
        # 환산지수 데이터
        cf_data = {}
        df_reval = processor.raw_data['df_sgr_reval']
        for htype in hospital_types:
            cf_data[htype] = {}
            for year in years:
                if year in df_reval.index and htype in df_reval.columns:
                    val = df_reval.loc[year, htype]
                    cf_data[htype][str(year)] = clean_val(val)
                else:
                    cf_data[htype][str(year)] = None
        
        # 건보대상자수
        # 건보대상자수
        pop_data = {}
        df_pop = processor.raw_data['df_pop']
        # Check available columns
        col_basic = '건보대상자수'
        col_aged = '건보_고령화반영후(대상자수)'
        
        for year in years:
            pop_data[str(year)] = {'basic': None, 'aged': None}
            if year in df_pop.index:
                if col_basic in df_pop.columns:
                    pop_data[str(year)]['basic'] = clean_val(df_pop.loc[year, col_basic])
                if col_aged in df_pop.columns:
                    pop_data[str(year)]['aged'] = clean_val(df_pop.loc[year, col_aged])
        
        # GDP 데이터
        gdp_data = {}
        df_gdp = processor.raw_data['df_gdp']
        for year in years:
            gdp_data[str(year)] = {}
            if year in df_gdp.index:
                for col in ['실질GDP', '영안인구']:
                    if col in df_gdp.columns:
                        val = df_gdp.loc[year, col]
                        gdp_data[str(year)][col] = clean_val(val)
        
        # 법과제도 데이터
        law_data = {}
        df_law = processor.raw_data['df_sgr_law']
        for htype in hospital_types:
            law_data[htype] = {}
            for year in years:
                if year in df_law.index and htype in df_law.columns:
                    val = df_law.loc[year, htype]
                    law_data[htype][str(year)] = clean_val(val)
                else:
                    law_data[htype][str(year)] = None
        
        # 상대가치변화
        rv_data = {}
        df_rv = processor.raw_data['df_rel_value']
        for htype in hospital_types:
            rv_data[htype] = {}
            for year in years:
                if year in df_rv.index and htype in df_rv.columns:
                    val = df_rv.loc[year, htype]
                    rv_data[htype][str(year)] = clean_val(val)
                else:
                    rv_data[htype][str(year)] = None
                    
        # 종별비용구조 (Weights)
        weights_data = {}
        df_w = processor.raw_data['df_weights']
        for htype in hospital_types:
            weights_data[htype] = {}
            for col in ['인건비', '관리비', '재료비']:
                 if htype in df_w.index and col in df_w.columns:
                     val = df_w.loc[htype, col]
                     weights_data[htype][col] = clean_val(val)
                 else:
                     weights_data[htype][col] = None

        # 급여율 (Benefit Rates)
        rate_data = {}
        df_rate = processor.raw_data.get('df_rate_py', pd.DataFrame())
        if not df_rate.empty:
            for htype in hospital_types:
                rate_data[htype] = {}
                for year in years:
                    if year in df_rate.index and htype in df_rate.columns:
                        val = df_rate.loc[year, htype]
                        rate_data[htype][str(year)] = clean_val(val)
                    else:
                        rate_data[htype][str(year)] = None

        # [NEW] 수가계약 (Contract)
        contract_data = {}
        df_contract = processor.raw_data.get('df_contract', pd.DataFrame())
        if not df_contract.empty:
            for year in years:
                contract_data[str(year)] = {}
                if year in df_contract.index:
                    for col in ['인상율_전체', '추가소요재정_전체']:
                        if col in df_contract.columns:
                            val = df_contract.loc[year, col]
                            contract_data[str(year)][col] = clean_val(val)
                        else:
                            contract_data[str(year)][col] = None

        # [NEW] 건보 재정통계 (Finance)
        finance_data = {}
        df_finance = processor.raw_data.get('df_finance', pd.DataFrame())
        if not df_finance.empty:
            for year in years:
                finance_data[str(year)] = {}
                if year in df_finance.index:
                    for col in df_finance.columns:
                        val = df_finance.loc[year, col]
                        finance_data[str(year)][col] = clean_val(val)

        return jsonify({
            'mei': mei_data,
            'medical': medical_data,
            'cf': cf_data,
            'population': pop_data,
            'gdp': gdp_data,
            'law': law_data,
            'rv': rv_data,
            'weights': weights_data,
            'benefit_rate': rate_data,
            'contract': contract_data,
            'finance': finance_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/get_excel_raw_data')
@login_required
def get_excel_raw_data():
    """메모리에 로드된 데이터를 반환"""
    try:
        # [OPTIMIZATION] Removed forced reload to speed up view switching.
        # Data remains in sync via save_to_excel_file route logic.
        
        # 이미 로드된 DataProcessor의 데이터를 활용
        # 매핑: 한글 표시명 -> processor.raw_data 내부 키
        data_map = {
            '진료비_실제': 'df_expenditure',
            '종별비용구조': 'df_weights',
            '생산요소_물가': 'df_raw_mei_inf',
            '1인당GDP': 'df_gdp',
            '건보대상': 'df_pop',
            '연도별환산지수': 'df_sgr_reval',
            '법과제도': 'df_sgr_law',
            '상대가치변화': 'df_rel_value',
            '기관수': 'df_num',
            '수가계약결과': 'df_contract',
            '건보_재정통계': 'df_finance',
            '급여율': 'df_rate_py'
        }
        
        results = {}
        for kor_name, internal_key in data_map.items():
            if internal_key in processor.raw_data:
                df = processor.raw_data[internal_key]
                if df is None or df.empty:
                    continue
                
                # 데이터 정제 (NaN -> None 변환)
                # JSON 직렬화를 위해 float('nan')이나 np.nan을 None으로 변환. Infinity도 처리.
                # Must convert to object dtype to hold None instead of NaN
                df_temp = df.replace([np.inf, -np.inf], np.nan)
                
                # [USER FIX] RVS 시트의 경우 빈 컬럼(Unnamed) 제거
                if kor_name == '상대가치변화':
                    df_temp = df_temp.loc[:, ~df_temp.columns.astype(str).str.contains('^Unnamed')]

                df_clean = df_temp.astype(object).where(pd.notnull(df_temp), None)
                
                # Index를 컬럼으로 변환하여 함께 표시 (엑셀처럼)
                df_display = df_clean.reset_index()
                
                # 컬럼명 처리 (Index 이름이 없으면 '구분' 등으로 표시하거나 기본값 사용)
                cols = df_display.columns.tolist()
                cols = [(str(c) if c is not None else '') for c in cols] # 컬럼명 문자열 변환
                
                results[kor_name] = {
                    'headers': cols,
                    'rows': df_display.values.tolist()
                }
                
        return jsonify(results)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ----------------------------------------------------------------------
# AI 최적화 API 엔드포인트
# ----------------------------------------------------------------------

@app.route('/api/ai_simulation', methods=['GET'])
def api_ai_simulation():
    """AI 시뮬레이션 결과 반환 (k, j 파라미터 최적화)"""
    if not AI_MODULE_AVAILABLE:
        return jsonify({'error': 'AI module not available'}), 503
    
    try:
        from ai_optimizer import BudgetFunctionSimulator
        
        simulator = BudgetFunctionSimulator('SGR_data.xlsx')
        best_params, all_results = simulator.find_optimal_parameters(
            k_range=(1, 5),
            j_range=(1, 3),
            years=[2021, 2022, 2023, 2024, 2025]
        )
        
        if best_params is None:
            return jsonify({'error': 'Simulation failed'}), 500
        
        response = {
            'success': True,
            'optimal_k': int(best_params['k']),
            'optimal_j': int(best_params['j']),
            'mean_error': float(best_params['abs_mean_error']),
            'std_error': float(best_params.get('std_error', 0)),
            'year_errors': best_params.get('year_errors', {}),
            'all_combinations': all_results.to_dict('records') if all_results is not None else []
        }
        
        return jsonify(sanitize_data(response))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai_optimization', methods=['POST'])
def api_ai_optimization():
    """AI 최적화 결과 반환 (S1 MEI-평균 기준 연동)"""
    if not AI_MODULE_AVAILABLE:
        return jsonify({'error': 'AI module not available'}), 503
    
    try:
        data = request.json or {}
        target_year = int(data.get('year') or data.get('target_year', 2026))
        sgr_results = data.get('sgr_results')
        
        # S1 MEI-Average reference data lookup
        if not sgr_results and _cached_analysis:
            try:
                bulk = _cached_analysis.get('bulk_sgr', {})
                if 'scenario_adjustments' in bulk and target_year in bulk['scenario_adjustments']:
                    sgr_results = bulk['scenario_adjustments'][target_year].get('평균', {}).get('S1', {})
            except: pass
            
        if not sgr_results:
            # Fallback if cache not found or year out of range
            sgr_results = {'병원(계)': 1.96, '의원': 1.9, '치과(계)': 1.96, '한방(계)': 1.96, '약국': 2.8}
            
        # 통합 엔진 사용 (고성능)
        engine = AIOptimizationEngine(data_frames=processor.raw_data)
        results = engine.run_full_analysis(target_year=target_year, sgr_results=sgr_results)
        
        if results is None:
            return jsonify({'error': 'AI analysis failed'}), 500
        
        return jsonify(sanitize_data(results))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai_full_report', methods=['POST'])
@login_required
def api_ai_full_report():
    """AI 전체 분석 리포트 반환 (S1 MEI-평균 기준 연동)"""
    if not AI_MODULE_AVAILABLE:
        return jsonify({'error': 'AI module not available'}), 503
    
    try:
        data = request.json or {}
        target_year = int(data.get('target_year', 2026))
        sgr_results = data.get('sgr_results')
        
        # S1 MEI-Average reference data lookup
        if not sgr_results and _cached_analysis:
            try:
                bulk = _cached_analysis.get('bulk_sgr', {})
                if 'scenario_adjustments' in bulk and target_year in bulk['scenario_adjustments']:
                    sgr_results = bulk['scenario_adjustments'][target_year].get('평균', {}).get('S1', {})
            except: pass

        if not sgr_results:
            sgr_results = {'병원(계)': 1.96, '의원': 1.9, '치과(계)': 1.96, '한방(계)': 1.96, '약국': 2.8}

        engine = AIOptimizationEngine(data_frames=processor.raw_data)
        results = engine.run_full_analysis(target_year=target_year, sgr_results=sgr_results)
        
        if results is None:
            return jsonify({'error': 'AI analysis failed'}), 500
        
        return jsonify(sanitize_data(results))
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=True)